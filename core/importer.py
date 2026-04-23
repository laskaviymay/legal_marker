from __future__ import annotations

import hashlib
import html
import importlib
import json
import logging
import re
import sys
import zipfile
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree as ET

from .forms import build_card_forms
from .manual_registry import (
    ManualRegistry,
    ensure_manual_registry_workbook,
    load_manual_registry,
    sync_manual_registry_cards,
)
from .models import LegalEntity, STATUS_ORDER
from .normalizer import (
    COUNTRY_ALIASES,
    extract_aliases,
    normalize_key,
    normalize_text,
    significant_alias,
    split_aliases,
    strip_alias_fragments,
    unique_clean,
)

LOGGER = logging.getLogger(__name__)
WORD_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}

SOURCE_STATUS_MAP: dict[str, tuple[str, ...]] = {
    "foreign_agents": ("foreign_agent",),
    "undesirable_organizations": ("undesirable_organization",),
    "rosfinmonitoring": ("terrorist_extremist",),
    "extremist_materials": ("extremist_material",),
}


def builtin_entities() -> list[LegalEntity]:
    return [
        LegalEntity(
            id="builtin_extremist_aue",
            name="АУЕ",
            entity_type="organization",
            statuses=("extremist",),
            aliases=(
                "Арестантское уголовное единство",
                "Арестантский уклад един",
            ),
            source="builtins",
            metadata={
                "source_file": "Встроенная запись",
                "note": "Системная карточка для устойчивого поиска сокращения АУЕ.",
            },
        ),
        LegalEntity(
            id="builtin_terrorist_antiwar_committee",
            name="Russian antiwar committee",
            entity_type="organization",
            statuses=("terrorist_extremist",),
            aliases=(
                "Антивоенный комитет России",
                "АКР",
                "RAC",
            ),
            source="builtins",
            metadata={
                "source_file": "Встроенная запись",
                "note": "Дополняющая карточка для объединения статусов Антивоенного комитета России.",
            },
        ),
    ]


ORG_DESCRIPTOR_PREFIX_RE = re.compile(
    r"^(?:(?:НАЦИОНАЛИСТИЧЕСКАЯ|НАЦИОНАЛИСТИЧЕСКОЕ|НАЦИОНАЛИСТИЧЕСКИЙ|"
    r"ТЕРРОРИСТИЧЕСКАЯ|ТЕРРОРИСТИЧЕСКОЕ|ТЕРРОРИСТИЧЕСКИЙ|"
    r"ЭКСТРЕМИСТСКАЯ|ЭКСТРЕМИСТСКОЕ|ЭКСТРЕМИСТСКИЙ|"
    r"МЕЖДУНАРОДНАЯ|МЕЖДУНАРОДНОЕ|МЕЖДУНАРОДНЫЙ)\s+)+"
    r"(?:ОРГАНИЗАЦИЯ|ДВИЖЕНИЕ|ГРУППИРОВКА|ОБЪЕДИНЕНИЕ|ПРОЕКТ)\s+(.+)$",
    re.IGNORECASE,
)


def import_foreign_agents(path: Path) -> list[LegalEntity]:
    rows = _read_xlsx_rows(path)
    entities: list[LegalEntity] = []
    for row in rows[3:]:
        raw_name = _cell(row, 1)
        if not raw_name:
            continue
        entity_type = "person" if "физ" in _cell(row, 6).lower() else "organization"
        name = strip_alias_fragments(raw_name) if entity_type == "person" else _clean_organization_name(raw_name)
        aliases = tuple(
            alias
            for alias in _derived_aliases(
                raw_name,
                name,
                include_acronyms=entity_type == "organization",
                include_person_aliases=entity_type == "person",
            )
            if significant_alias(alias)
        )
        entities.append(
            LegalEntity(
                id=_stable_id("foreign_agent", name),
                name=name,
                entity_type=entity_type,
                statuses=("foreign_agent",),
                aliases=aliases,
                source="foreign_agents",
                metadata=_row_metadata(rows[2] if len(rows) > 2 else [], row, path),
            )
        )
    return entities


def import_undesirable_organizations(path: Path) -> list[LegalEntity]:
    rows = _read_xlsx_rows(path)
    entities: list[LegalEntity] = []
    for row in rows[3:]:
        status = _cell(row, 9).lower()
        if status and "исключ" in status:
            continue
        raw_name = _cell(row, 4)
        if not raw_name:
            continue
        name = _clean_organization_name(_remove_parenthetical_fragments(raw_name))
        aliases = tuple(alias for alias in _derived_aliases(raw_name, name, include_acronyms=True) if significant_alias(alias))
        entities.append(
            LegalEntity(
                id=_stable_id("undesirable_organization", name),
                name=name,
                entity_type="organization",
                statuses=("undesirable_organization",),
                aliases=aliases,
                source="undesirable_organizations",
                metadata=_row_metadata(rows[2] if len(rows) > 2 else [], row, path),
            )
        )
    return entities


def import_rosfinmonitoring_docx(path: Path) -> list[LegalEntity]:
    paragraphs = _read_docx_paragraphs(path)
    return _import_rosfinmonitoring_lines(paragraphs, path)


def import_rosfinmonitoring(path: Path) -> list[LegalEntity]:
    suffix = path.suffix.lower()
    if suffix == ".docx":
        return import_rosfinmonitoring_docx(path)
    if suffix in {".html", ".htm"}:
        return import_rosfinmonitoring_html(path)
    if suffix in {".xls", ".xlsx"}:
        return import_rosfinmonitoring_table(path)
    raise ValueError(f"Unsupported Rosfinmonitoring file format: {path.suffix}")


def import_rosfinmonitoring_table(path: Path) -> list[LegalEntity]:
    rows = _read_table_rows(path)
    lines = [_row_text(row) for row in rows]
    return _import_rosfinmonitoring_lines(lines, path)


def import_rosfinmonitoring_html(path: Path) -> list[LegalEntity]:
    lines = _read_html_lines(path)
    return _import_rosfinmonitoring_lines(lines, path)


def _import_rosfinmonitoring_lines(lines: list[str], path: Path) -> list[LegalEntity]:
    entities: list[LegalEntity] = []
    section = "unknown"
    for line_number, paragraph in enumerate(lines, start=1):
        if not paragraph:
            continue
        lowered = paragraph.lower()
        if _is_rosfinmonitoring_heading(lowered, "организации"):
            section = "organization"
            continue
        if _is_rosfinmonitoring_heading(lowered, "физические лица"):
            section = "person"
            continue
        if section not in {"organization", "person"}:
            continue
        entry = _strip_number(paragraph)
        if not entry:
            continue
        name, aliases = _parse_ros_entry(entry, section)
        if not name:
            continue
        entities.append(
            LegalEntity(
                id=_stable_id("terrorist_extremist", name),
                name=name,
                entity_type=section,
                statuses=("terrorist_extremist",),
                aliases=aliases,
                source="rosfinmonitoring",
                metadata={"source_file": str(path), "line_number": str(line_number)},
            )
        )
    return entities


def import_extremist_materials_docx(path: Path, limit: int | None = None) -> list[LegalEntity]:
    paragraphs = _read_docx_paragraphs(path)
    entities: list[LegalEntity] = []
    current_number = ""
    for paragraph in paragraphs[1:]:
        if re.fullmatch(r"\d+[\.)]?", paragraph):
            current_number = paragraph.strip(".")
            continue
        if not current_number:
            continue
        title = paragraph.split("решение вынесено", 1)[0].strip(" ;")
        if not title:
            continue
        entities.append(
            LegalEntity(
                id=_stable_id("extremist_material", current_number, title),
                name=title[:300],
                entity_type="material",
                statuses=("extremist_material",),
                aliases=(),
                source="extremist_materials",
                metadata={"number": current_number, "source_file": str(path)},
            )
        )
        current_number = ""
        if limit is not None and len(entities) >= limit:
            break
    return entities


def build_database(
    foreign_agents_path: Path,
    undesirable_path: Path,
    rosfinmonitoring_path: Path,
    extremist_materials_path: Path | None,
    output_dir: Path,
) -> list[LegalEntity]:
    registry_path = _manual_registry_path(output_dir)
    ensure_manual_registry_workbook(registry_path)
    manual_registry = load_manual_registry(registry_path)
    entities: list[LegalEntity] = []
    entities.extend(import_foreign_agents(foreign_agents_path))
    entities.extend(import_undesirable_organizations(undesirable_path))
    entities.extend(import_rosfinmonitoring(rosfinmonitoring_path))
    if extremist_materials_path and extremist_materials_path.exists():
        entities.extend(import_extremist_materials_docx(extremist_materials_path, limit=None))
    entities.extend(builtin_entities())
    merged = _apply_manual_registry_overrides(merge_entities(entities), manual_registry)
    synced_registry = sync_manual_registry_cards(registry_path, merged)
    write_database(merged, output_dir, manual_registry=synced_registry)
    return merged


def import_source_file(source: str, path: Path) -> list[LegalEntity]:
    if source == "foreign_agents":
        return import_foreign_agents(path)
    if source == "undesirable_organizations":
        return import_undesirable_organizations(path)
    if source == "rosfinmonitoring":
        return import_rosfinmonitoring(path)
    if source == "extremist_materials":
        return import_extremist_materials_docx(path, limit=None)
    raise ValueError(f"Unknown source type: {source}")


def update_database_source(source: str, source_path: Path, output_dir: Path) -> list[LegalEntity]:
    if source not in SOURCE_STATUS_MAP:
        raise ValueError(f"Unknown source type: {source}")
    registry_path = _manual_registry_path(output_dir)
    ensure_manual_registry_workbook(registry_path)
    manual_registry = load_manual_registry(registry_path)
    existing = load_database(output_dir) if (output_dir / "agents.json").exists() else []
    retained = [
        retained_entity
        for entity in existing
        if (retained_entity := _without_source(entity, source)) is not None
    ]
    merged = _apply_manual_registry_overrides(
        merge_entities(retained + import_source_file(source, source_path) + builtin_entities()),
        manual_registry,
    )
    synced_registry = sync_manual_registry_cards(registry_path, merged)
    write_database(merged, output_dir, manual_registry=synced_registry)
    return merged


def merge_entities(entities: list[LegalEntity]) -> list[LegalEntity]:
    parent = list(range(len(entities)))

    def find(index: int) -> int:
        while parent[index] != index:
            parent[index] = parent[parent[index]]
            index = parent[index]
        return index

    def union(left: int, right: int) -> None:
        left_root, right_root = find(left), find(right)
        if left_root != right_root:
            parent[right_root] = left_root

    by_surface: dict[str, int] = {}
    for index, entity in enumerate(entities):
        name_key = normalize_key(entity.name)
        previous_name = by_surface.get(f"name:{name_key}")
        if name_key and previous_name is not None:
            union(previous_name, index)
        elif name_key:
            by_surface[f"name:{name_key}"] = index
        for surface in _merge_surfaces(entity):
            previous = by_surface.get(surface)
            if previous is None:
                by_surface[surface] = index
                continue
            if entities[previous].source != entity.source:
                union(previous, index)

    groups: dict[int, LegalEntity] = {}
    for index, entity in enumerate(entities):
        key = find(index)
        prepared = _prepare_merged_entity(entity)
        if key in groups:
            groups[key] = groups[key].merged_with(prepared)
        else:
            groups[key] = prepared
    return sorted(groups.values(), key=lambda item: (item.entity_type, item.name))


def _without_source(entity: LegalEntity, source: str) -> LegalEntity | None:
    sources = tuple(part.strip() for part in entity.source.split(";") if part.strip())
    if source not in sources:
        return entity
    remaining_sources = tuple(part for part in sources if part != source)
    removed_statuses = set(SOURCE_STATUS_MAP[source])
    remaining_statuses = tuple(status for status in entity.statuses if status not in removed_statuses)
    if not remaining_sources or not remaining_statuses:
        return None
    return LegalEntity(
        id=entity.id,
        name=entity.name,
        entity_type=entity.entity_type,
        statuses=remaining_statuses,
        aliases=entity.aliases,
        source="; ".join(remaining_sources),
        metadata=entity.metadata,
    )


def _prepare_merged_entity(entity: LegalEntity) -> LegalEntity:
    ordered_statuses = tuple(status for status in STATUS_ORDER if status in entity.statuses)
    return LegalEntity(
        id=entity.id,
        name=entity.name,
        entity_type=entity.entity_type,
        statuses=ordered_statuses or entity.statuses,
        aliases=unique_clean(entity.aliases),
        source=entity.source,
        metadata=entity.metadata,
    )


def _merge_surfaces(entity: LegalEntity) -> tuple[str, ...]:
    surfaces = []
    for surface in (entity.name, *entity.aliases):
        key = normalize_key(surface)
        if key and significant_alias(surface):
            surfaces.append(key)
    return tuple(dict.fromkeys(surfaces))


def write_database(
    entities: list[LegalEntity],
    output_dir: Path,
    manual_registry: ManualRegistry | None = None,
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    agents_path = output_dir / "agents.json"
    aliases_path = output_dir / "aliases.json"
    forms_path = output_dir / "forms.json"
    sources_path = output_dir / "sources.json"
    manual_registry = manual_registry or load_manual_registry(_manual_registry_path(output_dir))

    forms_by_entity = _runtime_forms_by_entity(entities, manual_registry)
    persisted_entities = [_entity_with_runtime_forms(entity, forms_by_entity.get(entity.id, ())) for entity in entities]

    agents_path.write_text(
        json.dumps([entity.to_dict() for entity in persisted_entities], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    alias_rows = []
    form_rows = []
    for entity in persisted_entities:
        for alias in entity.aliases:
            alias_rows.append({"entity_id": entity.id, "alias": alias})
        for form in forms_by_entity.get(entity.id, ()):
            form_rows.append({"entity_id": entity.id, "text": form.text, "source_type": form.source_type})
    aliases_path.write_text(json.dumps(alias_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    forms_path.write_text(json.dumps(form_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    sources_path.write_text(
        json.dumps(
            {
                "entity_count": len(persisted_entities),
                "sources": sorted(set(filter(None, (entity.source for entity in persisted_entities)))),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def load_database(data_dir: Path) -> list[LegalEntity]:
    path = data_dir / "agents.json"
    if not path.exists():
        raise FileNotFoundError(f"Database file not found: {path}")
    entities = [LegalEntity.from_dict(item) for item in json.loads(path.read_text(encoding="utf-8"))]
    forms_path = data_dir / "forms.json"
    if not forms_path.exists():
        return entities
    forms_by_entity: dict[str, list[str]] = {}
    for item in json.loads(forms_path.read_text(encoding="utf-8")):
        entity_id = str(item.get("entity_id", ""))
        text = str(item.get("text", "")).strip()
        if entity_id and text:
            forms_by_entity.setdefault(entity_id, []).append(text)
    loaded: list[LegalEntity] = []
    for entity in entities:
        runtime_forms = tuple(dict.fromkeys(forms_by_entity.get(entity.id, ())))
        if not runtime_forms:
            loaded.append(entity)
            continue
        metadata = dict(entity.metadata)
        metadata["runtime_forms"] = runtime_forms
        loaded.append(
            LegalEntity(
                id=entity.id,
                name=entity.name,
                entity_type=entity.entity_type,
                statuses=entity.statuses,
                aliases=entity.aliases,
                source=entity.source,
                metadata=metadata,
            )
        )
    return loaded


def _manual_registry_path(output_dir: Path) -> Path:
    return output_dir / "manual_registry.xlsx"


def _apply_manual_registry_overrides(
    entities: list[LegalEntity],
    manual_registry: ManualRegistry,
) -> list[LegalEntity]:
    cards_by_id = {row.card_id: row for row in manual_registry.cards if row.card_id}
    alias_rows_by_id: dict[str, list[str]] = {}
    for row in manual_registry.manual_aliases:
        if row.card_id and row.alias:
            alias_rows_by_id.setdefault(row.card_id, []).append(row.alias)
    updated: list[LegalEntity] = []
    for entity in entities:
        card = cards_by_id.get(entity.id)
        aliases = entity.aliases + tuple(alias_rows_by_id.get(entity.id, ()))
        statuses = entity.statuses
        metadata = dict(entity.metadata)
        if card and card.manual_statuses:
            manual_statuses = tuple(
                status.strip() for status in re.split(r"\s*[;,]\s*", card.manual_statuses) if status.strip()
            )
            statuses = tuple(status for status in STATUS_ORDER if status in set(statuses + manual_statuses))
            if card.notes:
                metadata["manual_notes"] = card.notes
        updated.append(
            LegalEntity(
                id=entity.id,
                name=entity.name,
                entity_type=entity.entity_type,
                statuses=statuses,
                aliases=unique_clean(aliases),
                source=entity.source,
                metadata=metadata,
            )
        )
    return updated


def _runtime_forms_by_entity(
    entities: list[LegalEntity],
    manual_registry: ManualRegistry,
) -> dict[str, tuple]:
    manual_aliases_by_id: dict[str, list[str]] = {}
    for row in manual_registry.manual_aliases:
        if row.card_id and row.alias:
            manual_aliases_by_id.setdefault(row.card_id, []).append(row.alias)
    manual_forms_by_id: dict[str, list[str]] = {}
    for row in manual_registry.manual_forms:
        if row.card_id and row.form:
            manual_forms_by_id.setdefault(row.card_id, []).append(row.form)
    disabled_forms_by_id: dict[str, list[str]] = {}
    for row in manual_registry.disabled_auto_forms:
        if row.card_id and row.form:
            disabled_forms_by_id.setdefault(row.card_id, []).append(row.form)
    return {
        entity.id: build_card_forms(
            entity=entity,
            manual_aliases=tuple(manual_aliases_by_id.get(entity.id, ())),
            manual_forms=tuple(manual_forms_by_id.get(entity.id, ())),
            disabled_auto_forms=tuple(disabled_forms_by_id.get(entity.id, ())),
        )
        for entity in entities
    }


def _entity_with_runtime_forms(entity: LegalEntity, runtime_forms: tuple) -> LegalEntity:
    metadata = dict(entity.metadata)
    metadata["runtime_forms"] = tuple(form.text for form in runtime_forms)
    return LegalEntity(
        id=entity.id,
        name=entity.name,
        entity_type=entity.entity_type,
        statuses=entity.statuses,
        aliases=entity.aliases,
        source=entity.source,
        metadata=metadata,
    )


def _read_xlsx_rows(path: Path) -> list[list[object]]:
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required to import xlsx registries") from exc
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = openpyxl.load_workbook(BytesIO(path.read_bytes()), read_only=True, data_only=True)
    sheet = workbook.active
    sheet.reset_dimensions()
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def _read_xls_rows(path: Path) -> list[list[object]]:
    xlrd = _load_xlrd()
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = xlrd.open_workbook(str(path))
    sheet = workbook.sheet_by_index(0)
    return [
        [_xls_cell_value(sheet.cell(row_index, column_index)) for column_index in range(sheet.ncols)]
        for row_index in range(sheet.nrows)
    ]


def _read_table_rows(path: Path) -> list[list[object]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx_rows(path)
    if suffix == ".xls":
        return _read_xls_rows(path)
    raise ValueError(f"Unsupported spreadsheet format: {path.suffix}")


def _load_xlrd():
    try:
        import xlrd
    except ModuleNotFoundError as exc:
        missing_error: Exception | None = exc
    else:
        if hasattr(xlrd, "open_workbook"):
            return xlrd
        missing_error = RuntimeError("Installed xlrd package is incomplete.")

    project_dir = Path(__file__).resolve().parents[1]
    for wheel in sorted((project_dir / ".wheels").glob("xlrd-*.whl")):
        if str(wheel) not in sys.path:
            sys.path.insert(0, str(wheel))
        sys.modules.pop("xlrd", None)
        xlrd = importlib.import_module("xlrd")
        if hasattr(xlrd, "open_workbook"):
            return xlrd

    raise RuntimeError("xlrd>=2.0.1 is required to import .xls Rosfinmonitoring registries.") from missing_error


def _xls_cell_value(cell) -> object:
    value = cell.value
    if getattr(cell, "ctype", None) == 2 and isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _row_text(row: list[object]) -> str:
    return " ".join(_cell(row, index) for index in range(len(row)) if _cell(row, index)).strip()


def _read_docx_paragraphs(path: Path) -> list[str]:
    if not path.exists():
        raise FileNotFoundError(path)
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml")
    root = ET.fromstring(xml)
    paragraphs: list[str] = []
    for paragraph in root.findall(".//w:p", WORD_NS):
        text = "".join(node.text or "" for node in paragraph.findall(".//w:t", WORD_NS))
        text = html.unescape(re.sub(r"\s+", " ", text).strip())
        if text:
            paragraphs.append(text)
    return paragraphs


def _read_html_lines(path: Path) -> list[str]:
    if not path.exists():
        raise FileNotFoundError(path)
    content = path.read_text(encoding="utf-8", errors="ignore")
    content = re.sub(r"(?is)<script[^>]*>.*?</script>", " ", content)
    content = re.sub(r"(?is)<style[^>]*>.*?</style>", " ", content)
    content = re.sub(r"(?i)<br\s*/?>", "\n", content)
    content = re.sub(r"(?i)</(?:p|div|li|tr|td|th|h[1-6]|section|article)>", "\n", content)
    content = re.sub(r"(?s)<[^>]+>", " ", content)
    content = html.unescape(content)
    return [
        cleaned
        for line in content.splitlines()
        if (cleaned := re.sub(r"\s+", " ", line).strip())
    ]


def _parse_ros_entry(entry: str, section: str) -> tuple[str, tuple[str, ...]]:
    entry = entry.strip(" ,;")
    if section == "person":
        name = re.split(r",|\d{2}\.\d{2}\.\d{4}| г\.р\.", entry, maxsplit=1)[0]
        return name.strip(" ,;"), ()
    name = strip_alias_fragments(entry)
    name = re.sub(r"\s*,\s*;?$", "", name).strip(" ,;")
    name = name.replace("*", "").strip()
    aliases = tuple(alias for alias in _derived_aliases(entry, name, include_acronyms=True) if significant_alias(alias))
    return name, aliases


def _derived_aliases(
    raw_value: str,
    cleaned_name: str,
    include_acronyms: bool = False,
    include_person_aliases: bool = False,
) -> tuple[str, ...]:
    aliases = list(extract_aliases(raw_value))
    aliases.extend(_expanded_alias_fragments(*aliases))
    aliases.extend(_short_brand_aliases(*aliases, cleaned_name))
    aliases.extend(_x_letter_variants(*aliases))
    comma_parts = split_aliases(cleaned_name)
    if len(comma_parts) > 1:
        aliases.extend(part for part in comma_parts if _comma_alias_candidate(part))
    if include_person_aliases:
        aliases.extend(_person_search_aliases(cleaned_name, *aliases))
    if include_acronyms:
        aliases.extend(_organization_short_aliases(cleaned_name))
        for alias in tuple(aliases):
            aliases.extend(_organization_short_aliases(alias))
        for surface in (cleaned_name, *aliases):
            acronym = _generated_acronym_alias(surface)
            if acronym:
                aliases.append(acronym)
    return tuple(alias for alias in unique_clean(aliases) if normalize_key(alias) != normalize_key(cleaned_name))


def _expanded_alias_fragments(*values: str) -> tuple[str, ...]:
    aliases: list[str] = []
    for value in values:
        aliases.extend(split_aliases(value))
        aliases.extend(extract_aliases(value))
        stripped = strip_alias_fragments(value)
        if normalize_key(stripped) != normalize_key(value):
            aliases.append(stripped)
            aliases.extend(split_aliases(stripped))
    return unique_clean(aliases)


def _person_search_aliases(name: str, *existing_aliases: str) -> tuple[str, ...]:
    aliases: list[str] = []
    tokens = _raw_word_tokens(name)
    if tokens and len(tokens[0]) >= 5:
        aliases.append(tokens[0])
    for surface in (*aliases, *existing_aliases):
        aliases.extend(_latin_person_aliases(surface))
    return unique_clean(aliases)


def _raw_word_tokens(value: str) -> list[str]:
    return re.findall(r"[A-Za-zА-Яа-яЁё]+(?:[-'][A-Za-zА-Яа-яЁё]+)*", value)


def _latin_person_aliases(value: str) -> tuple[str, ...]:
    tokens = _raw_word_tokens(value)
    aliases: list[str] = []
    for token in tokens:
        if not re.fullmatch(r"[А-Яа-яЁё-]+", token):
            continue
        transliterated = _transliterate_ru(token)
        if transliterated:
            aliases.append(transliterated)
            if "sht" in transliterated.lower():
                aliases.append(_preserve_initial_case(transliterated.lower().replace("sht", "st")))
    return unique_clean(aliases)


def _transliterate_ru(value: str) -> str:
    table = {
        "а": "a",
        "б": "b",
        "в": "v",
        "г": "g",
        "д": "d",
        "е": "e",
        "ё": "e",
        "ж": "zh",
        "з": "z",
        "и": "i",
        "й": "y",
        "к": "k",
        "л": "l",
        "м": "m",
        "н": "n",
        "о": "o",
        "п": "p",
        "р": "r",
        "с": "s",
        "т": "t",
        "у": "u",
        "ф": "f",
        "х": "kh",
        "ц": "ts",
        "ч": "ch",
        "ш": "sh",
        "щ": "shch",
        "ъ": "",
        "ы": "y",
        "ь": "",
        "э": "e",
        "ю": "yu",
        "я": "ya",
        "-": "-",
    }
    transliterated = "".join(table.get(char.lower(), "") for char in value)
    return _preserve_initial_case(transliterated)


def _preserve_initial_case(value: str) -> str:
    if not value:
        return ""
    return value[0].upper() + value[1:]


def _short_brand_aliases(*values: str) -> tuple[str, ...]:
    aliases: list[str] = []
    for value in values:
        for part in re.split(r"\s+[\u2012-\u2015–—-]\s+", value, maxsplit=1)[:1]:
            candidate = part.strip(" ,;")
            if candidate != value.strip(" ,;") and _short_brand_candidate(candidate):
                aliases.append(candidate)
    return unique_clean(aliases)


def _short_brand_candidate(value: str) -> bool:
    if not (2 <= len(value) <= 24):
        return False
    if not re.search(r"\d", value):
        return False
    return bool(re.fullmatch(r"[A-Za-zА-Яа-яЁё0-9 ._+-]+", value))


def _x_letter_variants(*values: str) -> tuple[str, ...]:
    aliases: list[str] = []
    for value in values:
        if not re.search(r"\d[хx]\d|\d[хx]|[хx]\d", value, re.IGNORECASE):
            continue
        aliases.extend((value.replace("х", "x").replace("Х", "X"), value.replace("x", "х").replace("X", "Х")))
    return tuple(alias for alias in unique_clean(aliases) if alias not in values)


def _generated_acronym_alias(value: str) -> str:
    stop_words = {
        "автономная",
        "ано",
        "межрегиональная",
        "некоммерческая",
        "но",
        "движение",
        "группировка",
        "националистическая",
        "националистическое",
        "националистический",
        "объединение",
        "организация",
        "проект",
        "общественная",
        "общественное",
        "общественный",
        "объединение",
        "организация",
    }
    tokens = [
        token
        for token in re.findall(r"[a-zа-яё0-9]+", normalize_text(value))
        if token not in stop_words and len(token) > 1
    ]
    if len(tokens) < 2:
        return ""
    acronym = "".join(token[0] for token in tokens).upper()
    if 2 <= len(acronym) <= 8:
        return acronym
    return ""


def _organization_short_aliases(value: str) -> tuple[str, ...]:
    aliases: list[str] = []
    current = value.strip(" ,;")
    for _ in range(2):
        match = ORG_DESCRIPTOR_PREFIX_RE.match(current)
        if not match:
            break
        current = match.group(1).strip(" ,;")
        if current and normalize_key(current) != normalize_key(value):
            aliases.append(current)
    return unique_clean(aliases)


def _comma_alias_candidate(value: str) -> bool:
    cleaned = value.strip(" ,;")
    compact = re.sub(r"[^A-Za-zА-Яа-яЁё0-9]+", "", cleaned)
    if not compact:
        return False
    return bool(re.fullmatch(r"[A-ZА-ЯЁ0-9]{2,12}", compact))


def _clean_organization_name(value: str) -> str:
    cleaned = _remove_trailing_country(value)
    cleaned = cleaned.strip(" ,;")
    quote_pairs = (("«", "»"), ('"', '"'), ("“", "”"))
    for left, right in quote_pairs:
        if cleaned.startswith(left) and cleaned.endswith(right):
            cleaned = cleaned[len(left) : -len(right)]
            break
    return cleaned.strip(" ,;")


def _remove_trailing_country(value: str) -> str:
    parts = split_aliases(value)
    if len(parts) > 1 and normalize_text(parts[-1]) in COUNTRY_ALIASES:
        return ", ".join(parts[:-1])
    return value


def _remove_parenthetical_fragments(value: str) -> str:
    return re.sub(r"\s*\([^)]*?\)", "", value).strip(" ,;")


def _strip_number(paragraph: str) -> str:
    return re.sub(r"^\s*\d+(?:\.0)?[\.)]?\s*", "", paragraph).strip()


def _is_rosfinmonitoring_heading(value: str, expected: str) -> bool:
    return normalize_text(value) == normalize_text(expected)


def _cell(row: list[object], index: int) -> str:
    if index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _row_metadata(headers: list[object], row: list[object], path: Path) -> dict[str, str]:
    metadata: dict[str, str] = {"source_file": str(path)}
    row_number = _cell(row, 0)
    if row_number:
        metadata["row_number"] = row_number
    for index, _ in enumerate(row):
        header = re.sub(r"\s+", " ", _cell(headers, index)).strip()
        value = _cell(row, index)
        if not header or not value:
            continue
        key = header
        if key in metadata:
            key = f"{header} ({index + 1})"
        metadata[key] = value
    return metadata


def _stable_id(*parts: str) -> str:
    digest = hashlib.sha1("|".join(parts).encode("utf-8")).hexdigest()
    return digest[:16]
