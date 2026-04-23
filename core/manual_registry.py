from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path


@dataclass(frozen=True)
class ManualCardRow:
    card_id: str
    canonical_name: str
    entity_type: str
    official_statuses: str = ""
    manual_statuses: str = ""
    source_keys: str = ""
    source_urls: str = ""
    is_active: str = "1"
    notes: str = ""


@dataclass(frozen=True)
class ManualAliasRow:
    card_id: str
    alias: str
    alias_type: str = "manual"
    comment: str = ""


@dataclass(frozen=True)
class ManualFormRow:
    card_id: str
    base_surface: str
    form: str
    form_type: str = "manual"
    comment: str = ""


@dataclass(frozen=True)
class DisabledAutoFormRow:
    card_id: str
    form: str
    reason: str = ""


@dataclass(frozen=True)
class ManualRegistry:
    cards: tuple[ManualCardRow, ...] = ()
    manual_aliases: tuple[ManualAliasRow, ...] = ()
    manual_forms: tuple[ManualFormRow, ...] = ()
    disabled_auto_forms: tuple[DisabledAutoFormRow, ...] = ()


SHEET_HEADERS: dict[str, list[str]] = {
    "cards": [
        "card_id",
        "canonical_name",
        "entity_type",
        "official_statuses",
        "manual_statuses",
        "source_keys",
        "source_urls",
        "is_active",
        "notes",
    ],
    "manual_aliases": ["card_id", "alias", "alias_type", "comment"],
    "manual_forms": ["card_id", "base_surface", "form", "form_type", "comment"],
    "disabled_auto_forms": ["card_id", "form", "reason"],
    "change_log": ["timestamp", "action", "card_id", "details"],
}


def ensure_manual_registry_workbook(path: Path) -> None:
    import openpyxl

    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "cards"
    first.append(SHEET_HEADERS["cards"])
    for sheet_name in ("manual_aliases", "manual_forms", "disabled_auto_forms", "change_log"):
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(SHEET_HEADERS[sheet_name])
    workbook.save(path)


def load_manual_registry(path: Path) -> ManualRegistry:
    import openpyxl

    ensure_manual_registry_workbook(path)
    workbook = openpyxl.load_workbook(path)
    return ManualRegistry(
        cards=_load_cards(workbook["cards"]),
        manual_aliases=_load_aliases(workbook["manual_aliases"]),
        manual_forms=_load_forms(workbook["manual_forms"]),
        disabled_auto_forms=_load_disabled(workbook["disabled_auto_forms"]),
    )


def save_manual_registry(path: Path, registry: ManualRegistry) -> None:
    import openpyxl

    ensure_manual_registry_workbook(path)
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "cards"
    _write_sheet(first, SHEET_HEADERS["cards"], registry.cards, _card_values)
    for sheet_name, rows, serializer in (
        ("manual_aliases", registry.manual_aliases, _alias_values),
        ("manual_forms", registry.manual_forms, _form_values),
        ("disabled_auto_forms", registry.disabled_auto_forms, _disabled_values),
        ("change_log", (), None),
    ):
        sheet = workbook.create_sheet(sheet_name)
        if serializer is None:
            sheet.append(SHEET_HEADERS[sheet_name])
        else:
            _write_sheet(sheet, SHEET_HEADERS[sheet_name], rows, serializer)
    workbook.save(path)


def sync_manual_registry_cards(path: Path, entities) -> ManualRegistry:
    existing = load_manual_registry(path)
    existing_by_id = {row.card_id: row for row in existing.cards if row.card_id}
    synced_cards: list[ManualCardRow] = []
    seen: set[str] = set()
    for entity in entities:
        previous = existing_by_id.get(entity.id)
        synced_cards.append(
            ManualCardRow(
                card_id=entity.id,
                canonical_name=entity.name,
                entity_type=entity.entity_type,
                official_statuses=";".join(entity.statuses),
                manual_statuses=previous.manual_statuses if previous else "",
                source_keys=entity.source,
                source_urls=str(entity.metadata.get("source_file", "")),
                is_active=previous.is_active if previous else "1",
                notes=previous.notes if previous else "",
            )
        )
        seen.add(entity.id)
    for row in existing.cards:
        if row.card_id not in seen:
            synced_cards.append(row)
    registry = ManualRegistry(
        cards=tuple(sorted(synced_cards, key=lambda item: (item.canonical_name.casefold(), item.card_id))),
        manual_aliases=existing.manual_aliases,
        manual_forms=existing.manual_forms,
        disabled_auto_forms=existing.disabled_auto_forms,
    )
    save_manual_registry(path, registry)
    return registry


def _sheet_rows(sheet) -> list[tuple[object, ...]]:
    rows = list(sheet.iter_rows(values_only=True))
    return rows[1:] if rows else []


def _load_cards(sheet) -> tuple[ManualCardRow, ...]:
    return tuple(
        ManualCardRow(*(str(value or "") for value in row[:9]))
        for row in _sheet_rows(sheet)
        if any(value not in ("", None) for value in row)
    )


def _load_aliases(sheet) -> tuple[ManualAliasRow, ...]:
    return tuple(
        ManualAliasRow(*(str(value or "") for value in row[:4]))
        for row in _sheet_rows(sheet)
        if any(value not in ("", None) for value in row)
    )


def _load_forms(sheet) -> tuple[ManualFormRow, ...]:
    return tuple(
        ManualFormRow(*(str(value or "") for value in row[:5]))
        for row in _sheet_rows(sheet)
        if any(value not in ("", None) for value in row)
    )


def _load_disabled(sheet) -> tuple[DisabledAutoFormRow, ...]:
    return tuple(
        DisabledAutoFormRow(*(str(value or "") for value in row[:3]))
        for row in _sheet_rows(sheet)
        if any(value not in ("", None) for value in row)
    )


def _write_sheet(sheet, headers: list[str], rows, serializer) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(serializer(row))


def _card_values(row: ManualCardRow) -> list[str]:
    return [
        row.card_id,
        row.canonical_name,
        row.entity_type,
        row.official_statuses,
        row.manual_statuses,
        row.source_keys,
        row.source_urls,
        row.is_active,
        row.notes,
    ]


def _alias_values(row: ManualAliasRow) -> list[str]:
    return [row.card_id, row.alias, row.alias_type, row.comment]


def _form_values(row: ManualFormRow) -> list[str]:
    return [row.card_id, row.base_surface, row.form, row.form_type, row.comment]


def _disabled_values(row: DisabledAutoFormRow) -> list[str]:
    return [row.card_id, row.form, row.reason]
